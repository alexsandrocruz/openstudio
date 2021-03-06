# -*- coding: utf-8 -*-

from general_helpers import max_string_length
from general_helpers import datestr_to_python
from general_helpers import set_form_id_and_get_submit_button

from openstudio.os_invoice import Invoice
from openstudio.os_invoices import Invoices
from openstudio.os_customer import Customer

from decimal import Decimal, ROUND_HALF_UP

import io
import weasyprint
import openpyxl


# @auth.requires(auth.has_membership(group_id='Admins') or \
#                auth.has_permission('read', 'invoices'))
# def list_invoices():
#     """
#         Lists invoices
#     """
#     session.invoices_invoice_payment_add_back = None
#
#     cuID = request.vars['cuID']
#     csID = request.vars['csID']
#     search_enabled = request.vars['search_enabled']
#     group_filter_enabled = request.vars['group_filter_enabled']
#
#     #print request.vars
#
#     # disable unused fields
#     db.invoices.id.readable = False
#     db.invoices.invoices_groups_id.readable = False
#     db.invoices.Footer.readable = False
#     db.invoices.Note.readable = False
#     db.invoices.Terms.readable = False
#
#     links = [ dict(header = T("Balance"),
#                    body   = list_invoices_get_balance),
#               list_invoices_get_buttons ]
#     left  = [ db.invoices_amounts.on(db.invoices_amounts.invoices_id == \
#                                      db.invoices.id) ]
#
#     fields = [ db.invoices.Status,
#                db.invoices.InvoiceID,
#                db.invoices.Description,
#                db.invoices.DateCreated,
#                db.invoices.DateDue,
#                db.invoices_amounts.TotalPrice,
#                db.invoices_amounts.VAT,
#                db.invoices_amounts.TotalPriceVAT ]
#
#     query = (db.invoices.id > 0)
#     # Status filter
#     query = list_invoices_get_status_query(query)
#     if search_enabled:
#         query = list_invoices_get_search_query(query)
#     if group_filter_enabled:
#         query = list_invoices_get_groups_query(query)
#
#     # General list, list for customer or list for subscription
#     if not cuID and not csID:
#         # list all invoices
#         db.invoices.auth_customer_id.readable = True
#         fields.insert(2, db.invoices.auth_customer_id)
#
#     if cuID:
#         query &= (db.invoices.auth_customer_id == cuID)
#     if csID:
#         query &= (db.invoices.customers_subscriptions_id == csID)
#         fields.insert(3, db.invoices.SubscriptionMonth)
#         fields.insert(4, db.invoices.SubscriptionYear)
#
#     delete_permission = auth.has_membership(group_id='Admins') or \
#                         auth.has_permission('delete', 'invoices')
#
#     headers = {'invoices.auth_customer_id':T("Customer")}
#
#     grid = SQLFORM.grid(query,
#         links=links,
#         left=left,
#         field_id=db.invoices.id,
#         fields=fields,
#         headers=headers,
#         create=False,
#         editable=False,
#         details=False,
#         searchable=False,
#         deletable=delete_permission,
#         csv=False,
#         #maxtextlengths=maxtextlengths,
#         orderby=~db.invoices.id,
#         ui = session.grid_ui)
#     grid.element('.web2py_counter', replace=None) # remove the counter
#     grid.elements('span[title=Delete]', replace=None) # remove text from delete button
#
#     form_search = ''
#     content = DIV()
#     if search_enabled:
#         response.js = 'set_form_classes();'
#         if 'search' in request.vars:
#             session.invoices_list_invoices_search = request.vars['search']
#             # date_created_from = datestr_to_python(DATE_FORMAT, request.vars['date_created_from'])
#             # print type(date_created_from)
#             # print date_created_from
#             # session.invoices_list_invoices_date_created_from = date_created_from
#             try:
#                 date_created_from = datestr_to_python(DATE_FORMAT, request.vars['date_created_from'])
#                 session.invoices_list_invoices_date_created_from = date_created_from
#             except ValueError:
#                 session.invoices_list_invoices_date_created_from = None
#             try:
#                 date_created_until = datestr_to_python(DATE_FORMAT,  request.vars['date_created_until'])
#                 session.invoices_list_invoices_date_created_until = date_created_until
#             except ValueError:
#                 session.invoices_list_invoices_date_created_until = None
#             try:
#                 date_due_from = datestr_to_python(DATE_FORMAT,  request.vars['date_due_from'])
#                 session.invoices_list_invoices_date_due_from = date_due_from
#             except ValueError:
#                 session.invoices_list_invoices_date_due_from = None
#             try:
#                 date_due_until = datestr_to_python(DATE_FORMAT, request.vars['date_due_until'])
#                 session.invoices_list_invoices_date_due_until = date_due_until
#             except ValueError:
#                 session.invoices_list_invoices_date_due_until = None
#
#             #print locals()
#
#             keys = ['search', 'date_created_from', 'date_created_until', 'date_due_from', 'date_due_until']
#             for key in keys:
#                 try:
#                     del request.vars[key]
#                 except KeyError:
#                     pass
#
#             # redirect to update page
#             redirect(URL(vars=request.vars))
#
#         form_search = list_invoices_get_form_search()
#         content.append(form_search)
#
#     form_groups = ''
#     if group_filter_enabled:
#         if 'invoices_groups_id' in request.vars:
#             session.invoices_list_invoices_group = request.vars['invoices_groups_id']
#
#             try:
#                 del request.vars['invoices_groups_id']
#             except KeyError:
#                 pass
#
#             # redirect to update page
#             redirect(URL(vars=request.vars))
#
#
#     # always add the grid
#     content.append(grid)
#
#     return dict(content=content)
#
#
# def list_invoices_get_form_search(var=None):
#     """
#         Returns search form for invoices page
#     """
# #    print type(session.invoices_list_invoices_date_created_from)
# #    print session.invoices_list_invoices_date_created_from
#
#     form = SQLFORM.factory(
#         Field('search',
#               default=session.invoices_list_invoices_search,
#               label=T('')),
#         Field('date_created_from', 'date',
#               requires=IS_EMPTY_OR(IS_DATE_IN_RANGE(format=DATE_FORMAT,
#                                                     minimum=datetime.date(1900, 1, 1),
#                                                     maximum=datetime.date(2999, 1, 1))),
#               #),
#               default=session.invoices_list_invoices_date_created_from),
#         Field('date_created_until', 'date',
#               requires=IS_EMPTY_OR(IS_DATE_IN_RANGE(format=DATE_FORMAT,
#                                                     minimum=datetime.date(1900, 1, 1),
#                                                     maximum=datetime.date(2999, 1, 1))),
#               #),
#               default=session.invoices_list_invoices_date_created_until),
#         Field('date_due_from', 'date',
#               requires=IS_EMPTY_OR(IS_DATE_IN_RANGE(format=DATE_FORMAT,
#                                                     minimum=datetime.date(1900, 1, 1),
#                                                     maximum=datetime.date(2999, 1, 1))),
#               #),
#               default=session.invoices_list_invoices_date_due_from),
#         Field('date_due_until', 'date',
#               requires=IS_EMPTY_OR(IS_DATE_IN_RANGE(format=DATE_FORMAT,
#                                                     minimum=datetime.date(1900, 1, 1),
#                                                     maximum=datetime.date(2999, 1, 1))),
#               #),
#               default=session.invoices_list_invoices_date_due_until),
#         submit_button = T('Go')
#     )
#
#     search = form.element('#no_table_search')
#     search['_class'] += ' margin-right'
#     search['_placeholder'] = T("Invoice #")
#
#
#     btn_clear = A(T("Clear"),
#                   _href=URL('invoices', 'list_invoices_clear_search',
#                             vars={'search_enabled':True}),
#                   _title=T("Clear search"),
#                   _class='btn btn-default',
#                   cid=request.cid)
#
#     form = DIV(
#         DIV(
#             form.custom.begin,
#             DIV(LABEL(T('Search')), BR(), form.custom.widget.search, _class='col-md-2'),
#             DIV(DIV(DIV(LABEL(T('Date from')), BR(), form.custom.widget.date_created_from, _class='col-md-2'),
#                     DIV(LABEL(T('Date until')), BR(), form.custom.widget.date_created_until, _class='col-md-2'),
#                     DIV(LABEL(T('Due from')), BR(), form.custom.widget.date_due_from, _class='col-md-2'),
#                     DIV(LABEL(T('Due until')), BR(), form.custom.widget.date_due_until, _class='col-md-2'),
#                     DIV(LABEL(T('Filter')), BR(), form.custom.submit, btn_clear, _class='col-md-3'),
#                     _class='row'),
#                 _class='col-md-8'),
#             form.custom.end,
#             _class='row'),
#     )
#
#     return form
#
#
@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'invoices'))
def list_invoices_clear_search():
    """
        Clears search for invoices page
    """
    session.invoices_list_invoices_search = None
    session.invoices_list_invoices_date_created_from = None
    session.invoices_list_invoices_date_created_until = None
    session.invoices_list_invoices_date_due_from = None
    session.invoices_list_invoices_date_due_until = None

    redirect(URL('finance', 'invoices', vars=request.vars))


def edit_get_link_add_payment(iID):
    """
        Returns an button and modal to add a payment for an invoice
    """
    content = LOAD('invoices', 'payment_add', ajax=False, ajax_trap=True, extension='load',
                    vars={'iID':iID})

    invoice = db.invoices(iID)
    title = T('Add payment for invoice') + ' #' + invoice.InvoiceID

    button_text = os_gui.get_modal_button_icon('credit-card')

    form_id = 'form_payment_add_' + str(iID)

    result = os_gui.get_modal(button_text=button_text,
                              button_title=T("Add payment"),
                              modal_title=title,
                              modal_content=content,
                              modal_footer_content=os_gui.get_submit_button(form_id),
                              modal_class=form_id,
                              modal_id='modal_payment_add_' + str(iID),
                              button_class='btn-sm')

    return result


@auth.requires_login()
def edit():
    """
        Shows edit page for an invoice
        request.vars['iID'] is expected to be invoices.id
    """
    iID = request.vars['iID']
    invoice = Invoice(iID)
    response.title = T("Invoice") + ' ' + invoice.invoice.InvoiceID
    response.subtitle = ''
    response.view = 'general/only_content_no_box.html'

    session.invoices_payment_add_back = 'invoices_edit'

    modals = DIV()

    try:
        cuID = db.invoices_customers(invoices_id = iID).auth_customer_id
    except AttributeError:
        cuID = None

    # TODO update for invoices_items

    try:
        csID = db.invoices_customers_subscriptions(invoices_id = iID).customers_subscriptions_id
    except AttributeError:
        csID = None
    try:
        cmID = db.invoices_customers_memberships(invoices_id = iID).customers_memberships_id
    except AttributeError:
        cmID = None

    return_url = edit_get_back(cuID, csID, cmID)

    # Disable CustomerListName Field to prevent w2p expecting data (it's not in the form)
    db.invoices.CustomerListName.readable=False
    db.invoices.CustomerListName.writable=False

    crud.messages.submit_button = T("Save")
    crud.messages.record_updated = T("Saved")
    crud.settings.update_next = URL('invoices', 'edit', vars={'iID':iID})
    crud.settings.update_onaccept = [ edit_set_amounts ]
    crud.settings.update_deletable = False
    form = crud.update(db.invoices, iID)

    # Tie the elements together using the form html5 attribute.
    elements = form.elements('input, select, textarea')
    for element in elements:
        element['_form'] = "MainForm"

    submit = form.element('input[type=submit]')
    submit['_class'] = 'btn'

    textareas = form.elements('textarea')
    for textarea in textareas:
        textarea['_class'] += ' tmced'

    back = os_gui.get_button('back', return_url)
    modal_payments = edit_get_modal_payments(iID)
    modal_payment_add = edit_get_link_add_payment(iID)
    modals.append(modal_payments)
    modals.append(modal_payment_add['modal'])
    payments = edit_get_payments_button(iID)
    pdf = os_gui.get_button('print',
                                 URL('pdf', vars={'iID':iID}),
                                 title='PDF',
                                 btn_size='btn')
    tools = edit_get_tools(iID)


    header_tools = SPAN(pdf, DIV(form.custom.submit, _class="pull-right"), DIV(payments, _class='pull-right'), tools)

    items = DIV(LOAD('invoices', 'list_items.load',
                     ajax=True,
                     ajax_trap=True,
                     vars={'iID':iID},
                     content=os_gui.get_ajax_loader(
                             message=T("Loading items"))))

    form = DIV(
        DIV(
            XML('<form id="MainForm" action="#" enctype="multipart/form-data" method="post">'),
            form.custom.end,
            DIV(
                DIV(DIV(DIV(H3(form.custom.label.Description, _class='box-title'),
                            _class='box-header'),
                        DIV(form.custom.widget.Description,
                            _class='box-body'),
                        _class='box box-primary'),
                    _class='col-md-12'),
                DIV(edit_get_studio_info(),
                    _class='col-md-6'),
                DIV(edit_get_customer_info(invoice, form),
                    _class='col-md-6'),
                DIV(DIV(DIV(H3(T("Link to invoice in customer profile"), _class='box-title'),
                            _class='box-header'),
                        DIV(DIV(URL('profile', 'invoice', vars={'iID': iID}, scheme=True, host=True),
                                _class='well'),
                            SPAN(T("This is a direct link to the invoice in the profile of this customer."), BR(),
                                 T("Please note that only the customer linked to this invoice will be able to view it after logging in."),
                                 _class="text-muted"),
                            _class='box-body'),
                        _class='box box-primary'),
                    _class='col-md-12'),
                _class='col-md-10 no-padding-left'),

            # options container
            DIV(edit_get_amounts(invoice),
                DIV(DIV(H3(T('Options'), _class='box-title'),
                           _class='box-header'),
                        DIV(DIV(LABEL(form.custom.label.InvoiceID),
                                form.custom.widget.InvoiceID,
                                _class='form-group'),
                            DIV(LABEL(form.custom.label.DateCreated),
                                form.custom.widget.DateCreated,
                                _class='form-group'),
                            DIV(LABEL(form.custom.label.DateDue),
                                form.custom.widget.DateDue,
                                _class='form-group'),
                            DIV(LABEL(form.custom.label.Status),
                                form.custom.widget.Status,
                                _class='form-group'),
                            DIV(LABEL(form.custom.label.payment_methods_id),
                                form.custom.widget.payment_methods_id,
                                _class='form-group'),
                            DIV(LABEL(T("Last updated")), BR(),
                                represent_datetime(invoice.invoice.Updated_at)
                                ),
                            _class='box-body'),
                        _class='box box-primary'),
                _class='col-md-2 no-padding-left'),
            _class="row"
        ),

        DIV(
            DIV(DIV(DIV(H3(T('Items'), _class='box-title'),
                        _class='box-header'),
                    DIV(items, _class='box-body'),
                    _class='box box-primary'),
                _class='col-md-12'),
            DIV(DIV(DIV(H3(T('Terms and Conditions'), _class='box-title'),
                        _class='box-header'),
                    DIV(DIV(form.custom.widget.Terms),
                        _class='box-body'),
                    _class='box box-primary'),
                _class='col-md-12'),
            DIV(DIV(DIV(H3(T('Footer'), _class='box-title'),
                        _class='box-header'),
                    DIV(form.custom.widget.Footer,
                        _class='box-body'),
                    _class='box box-primary'),
                _class='col-md-12'),
        _class='row')
    )

    credit_invoice_for = ''
    if invoice.invoice.credit_invoice_for:
        original_invoice = Invoice(invoice.invoice.credit_invoice_for)

        credit_invoice_for = os_gui.get_alert('info',
             SPAN(T('This is a credit invoice for invoice'), ' ',
                  A(original_invoice.invoice.InvoiceID,
                    _href=URL('edit', vars={'iID':original_invoice.invoice.id}))))


    content = DIV(credit_invoice_for, form, modals)

    return dict(content=content,
                header_tools=header_tools,
                back=back)


def edit_get_tools(iID):
    """
        :param iID: db.invoices.id
        :return: tools dropdown for invoice
    """
    invoice_tools = []

    # teacher holidays
    permission = auth.has_membership(group_id='Admins') or \
                 auth.has_permission('create', 'invoices')

    if permission:
        link = A(os_gui.get_fa_icon('fa-ban'),
                 T("Cancel and create credit invoice"),
                 _href=URL('invoices', 'cancel_and_create_credit_invoice', vars={'iID': iID}),
                 _title=T('Cancel and create credit invoice'))
        invoice_tools.append(link)

        #Check if invoice is not for a subscription, teacher payment, claim, event, class, classcard, or membership
        query = (
            (db.invoices.id == iID) &
            (db.invoices.TeacherPayment == False) &
            (db.invoices.EmployeeClaim == False) &
            (db.invoices_items_classes_attendance.invoices_items_id == None) &
            (db.invoices_items_customers_subscriptions.invoices_items_id == None) &
            (db.invoices_items_customers_memberships.invoices_items_id == None) &
            (db.invoices_items_customers_classcards.invoices_items_id == None) &
            (db.invoices_items_workshops_products_customers.invoices_items_id == None)
        )
        left = [
            db.invoices_items.on(
                db.invoices_items.invoices_id ==
                db.invoices.id
            ),
            db.invoices_items_classes_attendance.on(
                db.invoices_items.id ==
                db.invoices_items_classes_attendance.invoices_items_id
            ),
            db.invoices_items_customers_subscriptions.on(
                db.invoices_items.id ==
                db.invoices_items_customers_subscriptions.invoices_items_id
            ),
            db.invoices_items_customers_memberships.on(
                db.invoices_items.id ==
                db.invoices_items_customers_memberships.invoices_items_id
            ),
            db.invoices_items_customers_classcards.on(
                db.invoices_items.id ==
                db.invoices_items_customers_classcards.invoices_items_id
            ),
            db.invoices_items_workshops_products_customers.on(
                db.invoices_items.id ==
                db.invoices_items_workshops_products_customers.invoices_items_id
            )
        ]

        row = db(query).select(db.invoices.ALL,
                               db.invoices_items_classes_attendance.invoices_items_id,
                               db.invoices_items_customers_memberships.invoices_items_id,
                               db.invoices_items_customers_subscriptions.invoices_items_id,
                               db.invoices_items_customers_classcards.invoices_items_id,
                               db.invoices_items_workshops_products_customers.invoices_items_id,
                               left=left
                               ).first()
        # print row
        if row:
            link = A(os_gui.get_fa_icon('fa-clone'),
                             T("Duplicate"),
                             _href=URL('invoices', 'duplicate_invoice', vars={'iID': iID}),
                             _title=T('Duplicate Invoice'))
            invoice_tools.append(link)


    # get menu
    tools = os_gui.get_dropdown_menu(invoice_tools,
                                     '',
                                     btn_size='',
                                     btn_icon='wrench',
                                     menu_class='pull-right')

    return tools


@auth.requires(auth.has_membership(group_id='Admins') or
               auth.has_permission('create', 'invoices'))
def duplicate_invoice():
    """
        Shows edit page for an invoice
        request.vars['iID'] is expected to be invoices.id
    """
    oldiID = request.vars['iID']
    oldinvoice = db(db.invoices.id == oldiID).select().first()

    iID= db.invoices.insert(
        invoices_groups_id= oldinvoice.invoices_groups_id,
        payment_methods_id = oldinvoice.payment_methods_id,
        Status = 'draft',
        CustomerCompany= oldinvoice.CustomerCompany,
        CustomerCompanyRegistration = oldinvoice.CustomerCompanyRegistration,
        CustomerCompanyTaxRegistration = oldinvoice.CustomerCompanyTaxRegistration,
        CustomerName = oldinvoice.CustomerName,
        CustomerListName = oldinvoice.CustomerListName,
        CustomerAddress = oldinvoice.CustomerAddress,
        Description= oldinvoice.Description,
        Terms = oldinvoice.Terms,
        Footer= oldinvoice.Footer,
        Note = oldinvoice.Note
    )
    query = (db.invoices_items.invoices_id == oldiID)
    rows = db(query).select()
    for row in rows:
        db.invoices_items.insert(
            invoices_id = iID,
            Sorting = row.Sorting,
            ProductName = row.ProductName,
            Description = row.Description,
            Quantity = row.Quantity,
            Price = row.Price,
            tax_rates_id = row.tax_rates_id,
            accounting_glaccounts_id = row.accounting_glaccounts_id,
            accounting_costcenters_id = row.accounting_costcenters_id,
        )

    query = (db.invoices_customers_orders.invoices_id == oldiID)
    rows = db(query).select()
    if rows:
        for row in rows:
            db.invoices_customers_orders.insert(
                invoices_id = iID,
                customers_orders_id = row.customers_orders_id
            )


    query = (db.invoices_customers.invoices_id == oldiID)
    row = db(query).select().first()
    if row:
        db.invoices_customers.insert(
            invoices_id = iID,
            auth_customer_id = row.auth_customer_id
        )

    session.flash = T("You are now editing the duplicated invoice")

    redirect(URL('edit', vars= {'iID': iID}))


def edit_set_amounts(form):
    """
    :param form: crud update form for db.invoices
    :return: None
    """
    iID = form.vars.id
    invoice = Invoice(iID)
    invoice.set_amounts()
    invoice.on_update()


def edit_get_amounts(invoice, formatted=True):
    """
    :param iID: db.invoices.is
    :return: returns box with total amounts & balance
    """
    # subtotal
    # tax
    # total
    # paid
    # balance
    amounts = invoice.get_amounts()

    subtotal = SPAN(SPAN(T('Subtotal'), _class='bold pull-left'),
                    SPAN(represent_decimal_as_amount(amounts.TotalPrice), _class='pull-right'))
    vat = SPAN(SPAN(T('VAT'), _class='bold pull-left'),
                    SPAN(represent_decimal_as_amount(amounts.VAT), _class='pull-right'))
    total = SPAN(SPAN(T('Total'), _class='bold pull-left'),
                    SPAN(represent_decimal_as_amount(amounts.TotalPriceVAT), _class='pull-right'))
    paid = SPAN(SPAN(T('Paid'), _class='bold pull-left'),
                    SPAN(represent_decimal_as_amount(amounts.Paid), _class='pull-right'))
    balance = SPAN(SPAN(T('Balance'), _class='bold pull-left'),
                    SPAN(represent_decimal_as_amount(amounts.Balance), _class='pull-right'))


    content = DIV(subtotal, BR(),
                  vat, BR(),
                  total, BR(),
                  paid, BR(),
                  balance)


    if formatted:
        box = DIV(DIV(H3(T('Balance'), _class='box-title'), _class='box-header'),
                  DIV(content, _class='box-body'),
                  _class='box box-primary')

        return box
    else:
        return content


def edit_get_back(cuID, csID=None, cmID=None):
    """
        Returns back link for invoice edit page
    """
    if session.invoices_edit_back == 'customers_invoices':
        url = URL('customers', 'invoices', vars={'cuID':cuID})
    if session.invoices_edit_back == 'customers_orders':
        url = URL('customers', 'orders', vars={'cuID':cuID})
    elif session.invoices_edit_back == 'customers_classcards':
        url = URL('customers', 'classcards', vars={'cuID':cuID})
    elif session.invoices_edit_back == 'customers_memberships':
        url = URL('customers', 'memberships', vars={'cuID':cuID})
    elif session.invoices_edit_back == 'customers_membership_invoices':
        url = URL('customers', 'membership_invoices', vars={'cuID':cuID,
                                                            'cmID':cmID})
    elif session.invoices_edit_back == 'customers_subscription_invoices':
        url = URL('customers', 'subscription_invoices', vars={'cuID':cuID,
                                                              'csID':session.invoices_edit_back_csID})
    elif session.invoices_edit_back == 'customer_events':
        url = URL('customers', 'events', vars={'cuID':cuID})
    elif session.invoices_edit_back == 'events_tickets_list_customers':
        url = URL('events', 'tickets_list_customers')
    elif session.invoices_edit_back == 'classes_attendance':
        url = URL('classes', 'attendance')
    elif session.invoices_edit_back == 'finance_invoices':
        url = URL('finance', 'invoices')
    elif session.invoices_edit_back == 'finance_batch_content':
        url = URL('finance', 'batch_content')
    elif session.invoices_edit_back == 'finance_employee_expenses_processed':
        url = URL('finance', 'employee_expenses_processed')
    elif session.invoices_edit_back == 'finance_teacher_payment_classes_processed':
        url = URL('finance', 'teacher_payment_classes', vars={'status': 'processed'})
    elif session.invoices_edit_back == 'reports_subscriptions_alt_prices':
        url = URL('reports', 'subscriptions_alt_prices')
    else:
        url = URL('customers', 'invoices', vars={'cuID':cuID})

    return url


def edit_get_modal_payments(iID):
    """
        Retuns payments div for an invoice
    """
    content = LOAD('invoices', 'invoice_payments.load', ajax=True,
                    vars={'iID':iID})

    title = T('Payments')

    button_text = T('Payments')

    result = os_gui.get_modal(button_text=button_text,
                              button_title=T("Payments for this invoice"),
                              modal_title=title,
                              modal_content=content,
                              modal_class='payments_' + str(iID),
                              modal_size='lg',
                              button_class='')

    return result['modal']


def edit_get_payments_button(iID):
    return XML("""
    <div class="btn-group">
            <button type="button" class="btn btn-default " title="Payments for this invoice" data-toggle="modal" data-target=".payments_{iID}"> Payments</button>
            <button type="button" class="btn btn-default dropdown-toggle" data-toggle="dropdown">
                <span class="caret"></span>
            </button>
            <ul class="dropdown-menu dropdown-menu-right" role="menu">
                <li><a href="#"  role="button" data-toggle="modal" data-target="#modal_payment_add_{iID}"><i class="fa fa-plus"></i>{title}</a></li>
            </ul>
        </div>""".format(iID=str(iID),
                         title=T("Add payment")))


def edit_get_studio_info(var=None):
    """
        Returns a div with address info for studio
    """
    try:
        organization = ORGANIZATIONS[ORGANIZATIONS['default']]

        company_name = organization['Name']
        company_address = organization['Address']
        company_email = organization['Email'] or ''
        company_phone = organization['Phone'] or ''
        company_registration = organization['Registration'] or ''
        company_tax_registration = organization['TaxRegistration'] or ''
    except KeyError:
        company_name = ''
        company_address = ''
        company_email = ''
        company_phone = ''
        company_registration = ''
        company_tax_registration = ''

    info = DIV(DIV(H3(T('From'), _class='box-title'),
                   _class='box-header'),
               DIV(B(company_name),
                   XML(company_address), BR(),
                   company_email, BR(),
                   company_phone, BR(),
                   company_registration, BR(),
                   company_tax_registration, BR(),
                   _class="box-body"),
               _class='box box-primary')

    return info


def edit_get_customer_info(invoice, form):
    """
        Returns a div with address info for a customer
    """
    form.element('#invoices_CustomerAddress')['_class'] = 'text'
    form.element('#invoices_CustomerAddress')['_height'] = '100'

    info = DIV(
        LABEL(form.custom.label.CustomerCompany),
        form.custom.widget.CustomerCompany,
        LABEL(form.custom.label.CustomerCompanyRegistration),
        form.custom.widget.CustomerCompanyRegistration,
        LABEL(form.custom.label.CustomerCompanyTaxRegistration),
        form.custom.widget.CustomerCompanyTaxRegistration,
        LABEL(form.custom.label.CustomerName),
        form.custom.widget.CustomerName,
        LABEL(form.custom.label.CustomerAddress),
        form.custom.widget.CustomerAddress,
        _class='box-body')

    link_customer = ''
    cuID = invoice.get_linked_customer_id()
    if cuID:
        link_customer = A(os_gui.get_fa_icon('fa-external-link'), ' ',
                          T('Customer profile'),
                          _href=URL('customers', 'edit', args=cuID))


    box = DIV(DIV(SPAN(link_customer, _class='pull-right'),
                  H3(T("To"), _class='box-title'),
                  _class='box-header'),
              info,
              _class='box box-primary')

    return box


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'invoices'))
def list_items():
    """
        Lists items for an invoice and shows an add form at the top of the
        page, intended to be used as LOAD component
        request.vars['iID'] is expected to be invoices.id
    """
    # call js for styling the form
    response.js = 'set_form_classes();'

    iID = request.vars['iID']

    # Ugly hack to reload div that contains the reload after deleting an item
    # otherwise after deleting, further submitting becomes impossible
    if request.extension == 'load':
        if 'reload_list' in request.vars:
            response.js += "$('#" + request.cid + "').get(0).reload()"

    form = list_items_get_form_add(iID)

    content = DIV(form.custom.begin)

    table = TABLE(THEAD(TR(
                     TH(_class='Sorting'),
                     TH(T('Product Name'), _class='ProductName'),
                     TH(T('Description'), _class='Description'),
                     TH(T('Qty'), _class='Quantity'),
                     TH(T('Price'), _class='Price'),
                     TH(T('Tax rate'), _class='TaxRate'),
                     # TH(SPAN(T('Subtotal'), _class='pull-right')),
                     # TH(SPAN(T('VAT'), _class='pull-right')),
                     TH(SPAN(T('Total'), _class='pull-right')),
                     TH(T('GLAccount'), _class='GLAccount'),
                     TH(T('Cost center'), _class='Costcenter'),
                     TH(),
                     _class='header')),
                  TR(TD(),
                     TD(form.custom.widget.ProductName),
                     TD(form.custom.widget.Description),
                     TD(form.custom.widget.Quantity),
                     TD(form.custom.widget.Price),
                     TD(form.custom.widget.tax_rates_id),
                     # TD(),
                     # TD(),
                     TD(),
                     TD(form.custom.widget.accounting_glaccounts_id),
                     TD(form.custom.widget.accounting_costcenters_id),
                     TD(DIV(form.custom.submit, _class='pull-right'))),
                  _class='table table-hover table-striped invoice-items small_font',
                  _id=iID) # set invoice id as table id, so we can pick it up from js when calling items_update_sorting() using ajaj

    left = [
        db.invoices_items_customers_subscriptions.on(
            db.invoices_items_customers_subscriptions.invoices_items_id ==
            db.invoices_items.id
        ),
        db.invoices_items_customers_classcards.on(
            db.invoices_items_customers_classcards.invoices_items_id ==
            db.invoices_items.id
        ),
        db.invoices_items_customers_memberships.on(
            db.invoices_items_customers_memberships.invoices_items_id ==
            db.invoices_items.id
        ),
        db.invoices_items_workshops_products_customers.on(
            db.invoices_items_workshops_products_customers.invoices_items_id ==
            db.invoices_items.id
        ),
        db.invoices_items_employee_claims.on(
            db.invoices_items_employee_claims.invoices_items_id ==
            db.invoices_items.id
        ),
        db.invoices_items_teachers_payment_classes.on(
            db.invoices_items_teachers_payment_classes.invoices_items_id ==
            db.invoices_items.id
        )
    ]

    query = (db.invoices_items.invoices_id == iID)
    rows = db(query).select(
        db.invoices_items.ALL,
        db.invoices_items_customers_subscriptions.id,
        db.invoices_items_customers_classcards.id,
        db.invoices_items_customers_memberships.id,
        db.invoices_items_workshops_products_customers.id,
        db.invoices_items_employee_claims.id,
        db.invoices_items_teachers_payment_classes.id,
        left=left,
        orderby=db.invoices_items.Sorting,
        groupby=db.invoices_items.id # workaround for update to 2019.02 where references were linked to invoice items instead of invoices.
        # For teacher payments, the only option was to link all payments of in an entire invoice to the first item.

    )

    for i, row in enumerate(rows):
        repr_row = list(rows[i:i+1].render())[0]

        btn_vars = {'iID':iID, 'iiID':row.invoices_items.id}
        btn_size = 'btn-xs'
        buttons = DIV(_class='btn-group btn-group-sm pull-right')
        permission = auth.has_membership(group_id='Admins') or \
                     auth.has_permission('update', 'invoices_items')
        if permission:
            # Edit button
            btn_edit = os_gui.get_button(
                'edit_notext',
                URL('item_edit', vars=btn_vars),
                cid=request.cid
            )
            buttons.append(btn_edit)
            # Duplicate button
            btn_duplicate = os_gui.get_button(
                'duplicate',
                URL('item_duplicate', vars=btn_vars),
                cid=request.cid
            )
            buttons.append(btn_duplicate)

            sort_handler = SPAN(_title=T("Click, hold and drag to change the order of items"),
                                _class='glyphicon glyphicon-option-vertical grey')
        else:
            sort_handler = ''

        permission = auth.has_membership(group_id='Admins') or \
                     auth.has_permission('delete', 'invoices_items')
        if permission and list_items_check_row_deletable(row):
            btn_delete = os_gui.get_button(
                'delete_notext',
                URL('item_delete_confirm', vars=btn_vars),
                btn_class='btn-danger',
                _class='pull-right',
                cid=request.cid
            )
        else:
            btn_delete = ''


        tr = TR(TD(sort_handler, _class='sort-handler movable'),
                TD(row.invoices_items.ProductName),
                TD(row.invoices_items.Description, _class='Description'),
                TD(row.invoices_items.Quantity),
                TD(SPAN(repr_row.invoices_items.Price, _class='pull-right')),
                TD(repr_row.invoices_items.tax_rates_id),
                TD(SPAN(repr_row.invoices_items.TotalPriceVAT,
                        _title=T("Subtotal: ") + CURRSYM + str(row.invoices_items.TotalPrice) + ' ' +\
                               T('VAT: ') + CURRSYM + str(row.invoices_items.VAT),
                        _class='pull-right')),
                TD(repr_row.invoices_items.accounting_glaccounts_id),
                TD(repr_row.invoices_items.accounting_costcenters_id),
                TD(btn_delete, buttons))

        table.append(tr)

    # Add totals
    invoice       = Invoice(iID)
    amounts_total = invoice.get_amounts()
    amounts_vat   = invoice.get_amounts_tax_rates(formatted=False)

    # try:
    tfoot = TFOOT()
    amounts = [ [ T('Sub total'), amounts_total.TotalPrice ] ]

    for tax_rate in amounts_vat:
        amounts.append( [ tax_rate['Name'], tax_rate['Amount']])

    amounts.append([T('Total')    , amounts_total.TotalPriceVAT ])

    for amount in amounts:
        tfoot.append(TR(TD(),
                        TD(),
                        TD(),
                        TD(),
                        TD(),
                        TD(amount[0], _class='bold'),
                        TD(SPAN(CURRSYM, ' ',
                                format(amount[1], '.2f'),
                                _class='bold pull-right')),
                        # TD(),
                        # TD(),
                        TD(),
                        TD(),
                        TD(),
                        ))
    table.append(tfoot)
    # except (AttributeError, ValueError):
        # pass # when an amount doesn't exist or for a new invoice when no amounts have been set yet

    content.append(table)
    content.append(form.custom.end)

    return dict(content=content)


def list_items_check_row_deletable(row):
    """
    Don't allow deleting of row when it's linked to one or more products
    :param row:
    :return:
    """
    deletable = True
    if row.invoices_items_customers_subscriptions.id:
        deletable = False
    elif row.invoices_items_customers_classcards.id:
        deletable = False
    elif row.invoices_items_customers_memberships.id:
        deletable = False
    elif row.invoices_items_workshops_products_customers.id:
        deletable = False
    elif row.invoices_items_employee_claims.id:
        deletable = False
    elif row.invoices_items_teachers_payment_classes.id:
        deletable = False

    return deletable


@auth.requires_login()
def item_edit():
    """
        Edit invoice item
    """
    # call js for styling the form
    response.js = 'set_form_classes();'

    iID = request.vars['iID']
    iiID = request.vars['iiID']

    return_url = item_edit_delete_get_return_url(iID)

    crud.messages.submit_button    = T('Save')
    crud.messages.record_updated   = T("Saved item")
    crud.settings.update_deletable = False
    crud.settings.update_next      = return_url
    crud.settings.update_onaccept  = [list_items_create_update_onaccept]
    form = crud.update(db.invoices_items, iiID)

    edit_item = SPAN(SPAN(T('You are now editing the item below'), BR(),
                     _class='bold red'),
                     SPAN(A(T('Cancel'),
                          _href=return_url,
                          cid=request.cid),
                          _class='pull-right'))
    content = DIV(edit_item,
                  form.custom.begin)

    table = TABLE(THEAD(TR(
                     TH(),
                     TH(T('GLAccount'), _class='GLAccount'),
                     TH(T('Product Name'), _class='ProductName'),
                     TH(T('Description'), _class='Description'),
                     TH(T('Quantity'), _class='Quantity'),
                     TH(T('Price'), _class='Price'),
                     TH(T('Tax rate'), _class='TaxRate'),
                     TH(T('GLAccount'), _class='GLAccount'),
                     TH(T('Cost center'), _class='Costcenter'),
                     TH(),
                     _class='header')),
                  TR(TD(),
                     TD(form.custom.widget.GLAccount),
                     TD(form.custom.widget.ProductName),
                     TD(form.custom.widget.Description),
                     TD(form.custom.widget.Quantity),
                     TD(form.custom.widget.Price),
                     TD(form.custom.widget.tax_rates_id),
                     TD(form.custom.widget.accounting_glaccounts_id),
                     TD(form.custom.widget.accounting_costcenters_id),
                     TD(DIV(form.custom.submit, _class='pull-right'))),
                  _class='table table-hover table-striped invoice-items small_font')

    content.append(table)
    content.append(form.custom.end)

    return dict(content=content)


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('update', 'invoices_items'))
def item_duplicate():
    """
    Duplicate invoice item
    """
    iID = request.vars['iID']
    iiID = request.vars['iiID']

    invoice = Invoice(iID)
    invoice.item_duplicate(iiID)

    redirect(item_edit_delete_get_return_url(iID))


def item_edit_delete_get_return_url(iID):
    """
        Return url for item_edit & item_delete
    """
    return URL('list_items', vars={'iID':iID, 'reload_list':True})


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('delete', 'invoices_items'))
def item_delete_confirm():
    """
        Confirm deleting item
    """
    iID = request.vars['iID']
    iiID = request.vars['iiID']

    item = db.invoices_items(iiID)

    question = SPAN(T('Do you really want to delete the following item?'),
                    _class='bold')
    yes = os_gui.get_button('noicon',
        URL('item_delete', vars={'iID':iID,
                                 'iiID':iiID}),
        title=T("Yes"),
        cid=request.cid,
        _class='btn',
        btn_size='')
    no = os_gui.get_button('noicon',
        URL('list_items', vars={'iID':iID}),
        title=T("No"),
        cid=request.cid,
        btn_class='btn-primary',
        btn_size='')
    buttons = DIV(yes,' ', no)


    content = DIV(H5(question, _class='red'),
              SPAN('"', item.ProductName, ' ', item.Description, '"'),
              BR(), BR(),
              DIV(buttons, _class='pull-right'))

    return content


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('delete', 'invoices_items'))
def item_delete():
    """
        Delete item from invoice
    """
    from openstudio.tools import OsTools

    db = current.db
    os_tools = OsTools()

    iID = request.vars['iID']
    iiID = request.vars['iiID']

    item = db.invoices_items(iiID)

    # Actually delete item
    query = (db.invoices_items.id == iiID)
    db(query).delete()

    ##
    # Update sorting so we have a consecutive order (important for rowSorter - Sorting should match with rowSorter index)
    # Move Sorting of all invoice items below the deleted one, up by 1
    ##
    if item.Sorting:
        query = (db.invoices_items.invoices_id == iID) & \
                (db.invoices_items.Sorting > item.Sorting)
        rows = db(query).select(db.invoices_items.ALL)
        for row in rows:
            row.Sorting = row.Sorting - 1
            row.update_record()

    # update invoice amounts (this also calls invoice.on_update() through set_amounts())
    list_items_set_invoice_amounts(iID)

    session.flash = T('Deleted item')

    redirect(item_edit_delete_get_return_url(iID))


def list_items_get_form_add(iID):
    """
        Returns add form for invoice items
    """
    db.invoices_items.invoices_id.default = iID

    invoice = Invoice(iID)
    db.invoices_items.Sorting.default = invoice.get_item_next_sort_nr()

    crud.messages.submit_button = T('Add')
    crud.messages.record_created = T("Added item")
    crud.settings.create_onaccept = [
        list_items_create_update_onaccept
    ]
    form = crud.create(db.invoices_items)

    return form


def list_items_create_update_onaccept(form):
    """
        Set the current amounts in the db after adding a new item
    """
    iiID = form.vars.id # invoice item
    iID = db.invoices_items(iiID).invoices_id # invoice
    list_items_set_invoice_amounts(iID)


def list_items_set_invoice_amounts(iID):
    """
        Set invoice amounts
        iiID is expected to be db.invoices_items.id
    """
    invoice = Invoice(iID)
    invoice.set_amounts() # this also calls invoice.on_update


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('update', 'invoices_items'))
def items_update_sorting():
    """
        Called as JSON, sets the sorting for invoice items in the db
    """
    # check if we're using json, if not, provide a fail message
    if not request.extension == 'json':
        return dict(status  = 'fail',
                    message = T("Please call this function as json"))

    status = 'OK'
    message = T('Changed sorting')

    iID       = request.vars['iID']
    old_index = int(request.vars['old_index'])
    new_index = int(request.vars['new_index'])

    query = (db.invoices_items.invoices_id == iID) & \
            (db.invoices_items.Sorting == old_index)
    changed_row = db(query).select(db.invoices_items.ALL).first()

    if old_index > new_index:
        query = (db.invoices_items.invoices_id == iID) & \
                (db.invoices_items.Sorting < old_index) & \
                (db.invoices_items.Sorting >= new_index)
        rows = db(query).select(db.invoices_items.ALL)
        for row in rows:
            row.Sorting = row.Sorting + 1
            row.update_record()
    else:
        query = (db.invoices_items.invoices_id == iID) & \
                (db.invoices_items.Sorting > old_index) & \
                (db.invoices_items.Sorting <= new_index)
        rows = db(query).select(db.invoices_items.ALL)
        for row in rows:
            row.Sorting = row.Sorting - 1
            row.update_record()


    changed_row.Sorting = new_index
    changed_row.update_record()

    invoice = Invoice(iID)
    invoice.on_update()


    return dict(status = status,
                message = message)


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'invoices_payments'))
def invoice_payments():
    """
        List payments for an invoice
    """
    iID = request.vars['iID']
    query = (db.invoices_payments.invoices_id == iID)
    rows = db(query).select(db.invoices_payments.ALL,
                            orderby=db.invoices_payments.PaymentDate)

    table = TABLE(_class='table table-striped table-hover')
    header = THEAD(TR(TH(T('Date')),
                      TH(T('Amount')),
                      TH(T('Method')),
                      TH(T('Note')),
                      TH(), # actions
                      ))
    table.append(header)

    for i, row in enumerate(rows):
        repr_row = list(rows[i:i+1].render())[0]

        onclick_del = "return confirm('" + \
                      T('Really delete this payment?') + "');"

        delete = os_gui.get_button('delete_notext',
                                   URL('payment_delete', vars={'ipID':row.id}),
                                   onclick=onclick_del,
                                   cid=request.cid,
                                   _class='pull-right')

        tr = TR(TD(repr_row.PaymentDate),
                TD(repr_row.Amount),
                TD(repr_row.payment_methods_id),
                TD(repr_row.Note),
                TD(delete))

        table.append(tr)


    content = table

    return dict(content=content)


@auth.requires_login()
def payment_add():
    """
        Add payments for an invoice
    """
    response.js = 'set_form_classes();'

    if not request.extension == 'load':
        response.view = 'general/only_content.html'
    response.title = T('Add payment')

    iID = request.vars['iID']
    invoice = Invoice(iID)

    response.subtitle = invoice.invoice.InvoiceID

    ## default values
    db.invoices_payments.invoices_id.default = iID
    # amount
    amounts = db.invoices_amounts(invoices_id=iID)
    try:
        db.invoices_payments.Amount.default = amounts.TotalPriceVAT.quantize(Decimal('.01'), rounding=ROUND_HALF_UP)
    except AttributeError:
        pass
    # payment method
    try:
        payment_info = db.customers_payment_info(
            auth_customer_id = invoice.auth_customer_id)
        default_method = payment_info.payment_methods_id
        db.invoices_payments.payment_methods_id.default = default_method
    except AttributeError:
        pass

    # if session.invoices_payment_add_back == 'invoices_invoice_payments':
    #     # Don't redirect client side here, stay in the modal on the invoice edit page
    #     create_next = URL('invoices', 'invoice_payments', vars={'iID':iID})
    # else:
    create_next = '/invoices/payment_add_redirect_oncreate?ipID=[id]'

    crud.messages.submit_button = T("Save")
    crud.messages.record_created = T("Saved")
    crud.settings.create_next = create_next
    crud.settings.create_onaccept = [ payment_add_update_status ]
    form = crud.create(db.invoices_payments)

    form_id = 'form_payment_add_' + str(iID)
    form_element = form.element('form')
    form['_id'] = form_id

    elements = form.elements('input, select, textarea')
    for element in elements:
        element['_form'] = form_id

    submit = form.element('input[type=submit]')
    back = os_gui.get_button('back',
                             payment_add_get_back(iID, invoice.get_linked_customer_id()))

    if request.extension == 'load':
        return dict(content=form)

    invoice = Invoice(iID)
    ic = db.invoices_customers(invoices_id = iID)
    customer = Customer(ic.auth_customer_id)

    description = ''
    if invoice.invoice.Description:
        description = '(' + invoice.invoice.Description + ')'

    invoice_info = DIV(
        H4(T('Invoice')),
        B(customer.row.display_name), BR(),
        invoice.invoice.InvoiceID, ' ', description, BR(),
        H4(T('Current balance')),
        edit_get_amounts(invoice, formatted=False)
    )

    content = DIV(DIV(form, _class='col-md-6'),
                  DIV(invoice_info, _class='col-md-3'))

    return dict(content=content, save=submit, back=back)


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('create', 'invoices_payments'))
def payment_add_redirect_oncreate():
    """
        Redirect to invoice list, from the client side, to leave the add modal
    """
    ipID = request.vars['ipID']
    payment = db.invoices_payments(ipID)
    invoice = Invoice(payment.invoices_id)
    cuID = invoice.get_linked_customer_id()
    iID  = invoice.invoice.id

    session.flash = T('Saved')

    redirect(payment_add_get_back(iID, cuID), client_side=True)


def payment_add_get_back(iID, cuID):
    """
        Returns redirect url to use after adding a payment
    """
    if session.invoices_payment_add_back == 'customers_invoices':
        url = URL('customers', 'invoices', vars={'cuID':cuID}, extension='')
    elif session.invoices_payment_add_back == 'customers_orders':
        url = URL('customers', 'orders', vars={'cuID':cuID}, extension='')
    elif session.invoices_payment_add_back == 'customers_classes_attendance':
        url = URL('customers', 'classes_attendance', vars={'cuID': cuID}, extension='')
    elif session.invoices_payment_add_back == 'customers_subscription_invoices':
        url = URL('customers', 'subscription_invoices',
                  vars={'cuID':cuID},
                  extension='')
    elif session.invoices_payment_add_back == 'customers_memberships':
        url = URL('customers', 'memberships',
                  vars={'cuID':cuID},
                  extension='')
    elif session.invoices_payment_add_back == 'customers_events':
        url = URL('customers', 'events',
                  vars={'cuID':cuID},
                  extension='')
    elif session.invoices_payment_add_back == 'classes_attendance':
        url = URL('classes', 'attendance', extension='')
    elif session.invoices_payment_add_back == 'finance_invoices':
        url = URL('finance', 'invoices', extension='')
    elif session.invoices_payment_add_back == \
                'events_tickets_list_customers':
        url = URL('events', 'tickets_list_customers', extension='')
    elif session.invoices_payment_add_back == 'customers_classcards':
        url = URL('customers', 'classcards', vars={'cuID':cuID}, extension='')
    elif session.invoices_payment_add_back == 'invoices_edit':
        url = URL('invoices', 'edit', vars={'iID':iID}, extension='')
    else: # catch all
        url = URL('customers', 'invoices', vars={'cuID':cuID}, extension='')

    return url


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('delete', 'invoices_payments'))
def payment_delete():
    """
        Delete invoice payment
    """
    ipID = request.vars['ipID']
    payment = db.invoices_payments(ipID)

    invoice = Invoice(payment.invoices_id)

    query = (db.invoices_payments.id == ipID)
    db(query).delete()

    # update invoice amounts
    invoice.is_paid()

    redirect(URL('invoice_payments', vars={'iID':payment.invoices_id}))


def payments_get_return_url(iID):
    """
        Return URL for payments
    """
    return URL('edit', vars={'iID':iID})


def payment_add_update_status(form):
    """
        Checks if the total of payments for an invoice
        >= invoices_amounts.TotalPriceVAT
        if so, set the invoice status to paid.
    """
    ipID = form.vars.id
    payment = db.invoices_payments(ipID)

    invoice = Invoice(payment.invoices_id)
    # check if the status should be changed
    invoice.is_paid()


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'invoices'))
def list_payments():
    """
        Lists all payments by a customer
    """
    cuID = request.vars['cuID']

    db.invoices.id.readable = False
    db.invoices.invoices_groups_id.readable = False
    db.invoices.Footer.readable = False
    db.invoices.Note.readable = False
    db.invoices.Terms.readable = False

    query = (db.invoices_customers.auth_customer_id == cuID)
    #links = [ list_invoices_get_links ]
    links = ''

    fields = [ db.invoices_payments.PaymentDate,
               db.invoices.DateDue,
               db.invoices.InvoiceID,
               db.invoices_amounts.TotalPriceVAT,
               db.invoices_payments.payment_methods_id,
               db.invoices_payments.Note ]

    left = [ db.invoices.on(db.invoices_payments.invoices_id ==
                            db.invoices.id),
             db.invoices_customers.on(db.invoices_customers.invoices_id ==
                                      db.invoices.id),
             db.invoices_amounts.on(db.invoices_payments.invoices_id ==
                                    db.invoices.id) ]

    delete_permission = auth.has_membership(group_id='Admins') or \
                        auth.has_permission('delete', 'invoices_payments')

    grid = SQLFORM.grid(query, links=links,
        fields=fields,
        left=left,
        #headers=headers,
        create=False,
        editable=False,
        details=False,
        searchable=False,
        deletable=delete_permission,
        csv=False,
        #maxtextlengths=maxtextlengths,
        orderby=~db.invoices_payments.PaymentDate,
        field_id=db.invoices_payments.id,
        ui = grid_ui)
    grid.element('.web2py_counter', replace=None) # remove the counter
    grid.elements('span[title=Delete]', replace=None) # remove text from delete button


    return dict(content=grid)


def pdf_template_get_logo(var=None):
    """
        Returns logo for pdf template
    """
    branding_logo = os.path.join(request.folder,
                                 'static',
                                 'plugin_os-branding',
                                 'logos',
                                 'branding_logo_invoices.png')
    if os.path.isfile(branding_logo):
        abs_url = '%s://%s/%s/%s' % (request.env.wsgi_url_scheme,
                                     request.env.http_host,
                                     'static',
                        'plugin_os-branding/logos/branding_logo_invoices.png')
        logo_img = IMG(_src=abs_url)

    else:
        logo_img = ''

    return logo_img


def pdf_template(iID):
    """
        Print friendly display of invoice
    """
    response.view = 'templates/invoices/default.html'
    iID = request.vars['iID']

    # get the invoice
    invoice = db.invoices(iID)
    # get invoice items
    query = (db.invoices_items.invoices_id == iID)
    items = db(query).select(db.invoices_items.ALL,
                             orderby=db.invoices_items.Sorting).render()
    # get customer info
    # query = (db.invoices_customers.invoices_id == invoice.id)
    # cuID = db(query).select(db.invoices_customers.auth_customer_id).first().auth_customer_id
    # print cuID
    #
    # query = (db.auth_user.id == cuID)
    # customer = db(query).select(db.auth_user.company,
    #                             db.auth_user.first_name,
    #                             db.auth_user.last_name,
    #                             db.auth_user.address,
    #                             db.auth_user.city,
    #                             db.auth_user.postcode,
    #                             db.auth_user.country,
    #                             db.auth_user.email,
    #                             db.auth_user.phone,
    #                             db.auth_user.mobile).first()

    # get amounts
    query = (db.invoices_amounts.invoices_id == iID)
    amounts = db(query).select(db.invoices_amounts.ALL)
    amounts = list(amounts[0:1].render())[0] # make the generator output stuff

    i = Invoice(iID)
    amounts_vat = i.get_amounts_tax_rates(formatted=True)

    # Company info
    try:
        organization = ORGANIZATIONS[ORGANIZATIONS['default']]

        company_name = organization['Name']
        company_address = organization['Address']
        company_email = organization['Email'] or ''
        company_phone = organization['Phone'] or ''
        company_registration = organization['Registration'] or ''
        company_tax_registration = organization['TaxRegistration'] or ''
        company_terms_conditions_url = organization['TermsConditionsURL']
    except KeyError:
        company_name = ''
        company_address = ''
        company_email = ''
        company_phone = ''
        company_registration = ''
        company_tax_registration = ''
        company_terms_conditions_url = ''

    studio = {}
    studio['name']         = company_name
    studio['address']      = XML(company_address)
    studio['email']        = company_email
    studio['phone']        = company_phone
    studio['registration'] = company_registration
    studio['tax_registration'] = company_tax_registration

    base_url = 'http%s://%s' % (request.is_https and 's' or '',
                                request.env.http_host)

    logo = pdf_template_get_logo()

    # print studio
    # print invoice
    # print items
    # print amounts
    # print amounts_vat
    # print customer
    # print logo
    # print base_url

    html = response.render('templates/invoices/default.html',
                           dict(invoice     = invoice,
                                items       = items,
                                amounts     = amounts,
                                amounts_vat = amounts_vat,
                                studio      = studio,
                                logo        = logo,
                                base_url    = base_url))

    return html


# No decorator here, permissions are checked inside the function
def pdf():
    """
        Converts a invoice to PDF
    """
    iID = request.vars['iID']

    invoice = Invoice(iID)

    permission  = ((auth.has_membership(group_id='Admins') or
                    auth.has_permission('read', 'invoices')) or
                   invoice.get_linked_customer_id() == auth.user.id)

    if not permission:
        return T("Not authorized")

    html = pdf_template(iID)

    fname = 'Invoice_' + invoice.invoice.InvoiceID + '.pdf'
    response.headers['Content-Type']='application/pdf'
    response.headers['Content-disposition']='attachment; filename=' + fname
    # return pyfpdf_from_html(html)

    stream = io.BytesIO()
    invoice = weasyprint.HTML(string=html).write_pdf(stream)

    return stream.getvalue()


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'invoices'))
def pdf_template_show():
    iID = request.vars['iID']

    return pdf_template(iID)


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'invoices'))
def export_invoices():
    """
        Shows a page which exports all invoices in a selected date range
    """
    response.title = T('Invoices')
    response.subtitle = T('Export invoices')

    response.view = 'general/tabs_menu.html'

    default_date = datetime.date(TODAY_LOCAL.year, TODAY_LOCAL.month, 1)
    from_date_default = default_date

    form = export_invoices_payments_get_form(default_date)

    # processing
    if form.process().accepted:
        response.flash = T("Form accepted")

        stream = export_invoices_get_export(
            form.vars.from_date,
            form.vars.until_date,
            form.vars.invoices_groups_id,
            form.vars.filetype,
            form.vars.include_subscriptions
        )

        if form.vars.filetype == 'excel':
            fname = "Invoices.xlsx"
            response.headers['Content-Type']='application/vnd.ms-excel'
            response.headers['Content-disposition']='attachment; filename=' + fname
        else:
            fname = "Invoices.csv"
            response.headers['Content-Type']='text/csv'
            response.headers['Content-disposition']='attachment; filename=' + fname

        return stream.getvalue()
        #response.stream(stream, attachment=True, filename=fname)


    # form.process() has to be called before creating a custom form
    # otherwise for hidden formkey fields used for CSRF protection aren't generated
    form_display = export_invoices_payments_get_form_display(form)
    content = DIV(form_display)

    menu = export_get_menu(request.function)

    back = os_gui.get_button('back', URL('finance', 'invoices'))

    return dict(content=content,
                back=back,
                menu=menu)


def export_invoices_get_export(from_date, until_date, invoices_groups_id, filetype='excel', include_subscriptions=True):
    """
        Invoices export
    """

    if filetype == 'excel':
        # create filestream
        stream = io.BytesIO()
        wb = openpyxl.workbook.Workbook(write_only=True)
        ws = wb.create_sheet(title='Invoices')
    else:
        import csv
        # create filestream
        stream = io.StringIO()
        csv_writer = csv.writer(stream, delimiter='\t')

    header = [
        'InvoiceID',
        'CustomerID',
        'Customer Name',
        'Business',
        'Date Created',
        'Date Due',
        'Status',
        'Description',
        'G/L Account',
        'Costcenter',
        'Item #',
        'Item Name',
        'Item Description',
        'Qty',
        'Price (each)',
        'Tax %',
        'Tax name',
        'Total excl. VAT',
        'VAT',
        'Total incl. VAT',
        'Payment Method',
        'School SubscriptionID',
        'School Subscription Name',
        'Subscription Year',
        'Subscription Month',
        'Payment date(s)'
    ]

    if filetype == 'excel':
        ws.append(header)
    else:
        csv_writer.writerow(header)


    where_query = "i.DateCreated >= '{from_date}'".format(from_date=from_date)

    if until_date:
        where_query += " AND i.DateCreated <= '{until_date}'".format(until_date=until_date)

    if invoices_groups_id:
        where_query += " AND i.invoices_groups_id = {invoices_groups_id}".format(
            invoices_groups_id=invoices_groups_id
        )

    if not include_subscriptions:
        where_query += " AND cs.id IS NULL"

    i = 0
    m = 1000
    while True:
        query = '''
        SELECT i.InvoiceID,
               au.id,
               au.display_name,
               CASE WHEN au.business = 'T' THEN "Yes" ELSE "No" END AS Business,
               i.DateCreated,
               i.DateDue,
               i.Status,
               i.Description,
               ii.accounting_glaccounts_id,
               ii.accounting_costcenters_id,
               ii.Sorting,
               ii.ProductName,
               ii.Description,
               ii.Quantity,
               ii.Price,
               tr.Percentage,
               tr.Name,
               ii.TotalPrice,
               ii.VAT,
               ii.TotalPriceVAT,
               pm.Name,
               ssu.id as ssuID,
               ssu.Name,
               i.SubscriptionYear,
               i.SubscriptionMonth,
               ip.payment_dates
        FROM invoices_items ii
        LEFT JOIN invoices i on ii.invoices_id = i.id
        LEFT JOIN (SELECT invoices_id, GROUP_CONCAT(PaymentDate SEPARATOR ', ') AS payment_dates
                   FROM invoices_payments GROUP BY invoices_id) ip ON ip.invoices_id = i.id
        LEFT JOIN invoices_customers ic ON ic.invoices_id = i.id
        LEFT JOIN invoices_items_customers_subscriptions iics ON iics.invoices_items_id = ii.id
        LEFT JOIN customers_subscriptions cs ON iics.customers_subscriptions_id = cs.id
        LEFT JOIN school_subscriptions ssu ON cs.school_subscriptions_id = ssu.id
        LEFT JOIN auth_user au ON ic.auth_customer_id = au.id
        LEFT JOIN tax_rates tr ON ii.tax_rates_id = tr.id
        LEFT JOIN payment_methods pm ON i.payment_methods_id = pm.id 
        WHERE {where_query} 
        ORDER BY i.InvoiceID, ii.Sorting
        LIMIT {limit} OFFSET {offset}
        '''.format(
            where_query=where_query,
            limit=m,
            offset=m * i
        )

        rows = db.executesql(query)

        for row in rows:
            unicode_list = []
            for item in row:
                try:
                    unicode_list.append(item.encode('utf-8'))
                except:
                    unicode_list.append(item)


            if filetype == 'excel':
                ws.append(unicode_list)
            else:
                csv_writer.writerow(unicode_list)


        # While loop control
        if len(rows) < m:
            break
        i += 1


    if filetype == 'excel':
        wb.save(stream)

    return stream


def export_invoices_payments_get_form(from_date_default, form_type='invoices'):
    """
        :param from_date_default: datetime.date
        :return: Invoices and payments export form
    """
    ig_query = (db.invoices_groups.Archived == False)

    form = SQLFORM.factory(
        Field('from_date', 'date',
            default=from_date_default,
            requires=IS_DATE_IN_RANGE(format=DATE_FORMAT,
                                 minimum=datetime.date(1900,1,1),
                                 maximum=datetime.date(2999,1,1))),
        Field('until_date', 'date',
            requires=IS_EMPTY_OR(IS_DATE_IN_RANGE(format=DATE_FORMAT,
                                 minimum=datetime.date(1900,1,1),
                                 maximum=datetime.date(2999,1,1)))),
        Field('invoices_groups_id', 'db.invoices_groups',
            requires=IS_EMPTY_OR(IS_IN_DB(db(ig_query),
                                          'invoices_groups.id',
                                          '%(Name)s',
                                          zero=T('All groups')))),
        Field('filetype',
              requires=IS_IN_SET([['excel', T('Excel (xlsx)')], ['tsv', T("Tab Separated Value (tsv)")]],
                                 zero=None)),
        Field('include_subscriptions', 'boolean',
              default=True),
        submit_button = T('Export')
    )

    return form


def export_invoices_payments_get_form_display(form, form_type='invoices'):
    """
    :param form: form returned by export_invoices_payments_get_form
    :return: display form
    """
    include_subscriptions = ''
    if form_type == 'invoices':
        include_subscriptions = DIV(
            DIV(LABEL(form.custom.widget.include_subscriptions, ' ',
                      T('Include subscriptions')),
                      _class='col-md-12'),
            _class='row'
        )


    form_display = DIV(
        form.custom.begin,
        DIV(DIV(LABEL(T('From date')), _class='col-md-2'),
            DIV(LABEL(T('Until date')), _class='col-md-2'),
            DIV(LABEL(T('Invoice group')), _class='col-md-2'),
            DIV(LABEL(T('Export filetype')), _class='col-md-2'),
            _class='row'),
        DIV(DIV(form.custom.widget.from_date, _class='col-md-2'),
            DIV(form.custom.widget.until_date, _class='col-md-2'),
            DIV(form.custom.widget.invoices_groups_id, _class='col-md-2'),
            DIV(form.custom.widget.filetype, _class='col-md-2'),
            _class='row'),
        include_subscriptions,
        form.custom.submit,
        form.custom.end,
    )

    return form_display


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'invoices'))
def export_payments():
    """
        Shows a page which exports all paymentsin a selected date range
    """
    response.title = T('Invoices')
    response.subtitle = T('Export payments')

    response.view = 'general/tabs_menu.html'

    today = datetime.date.today()
    default_date = datetime.date(today.year, today.month, 1)

    default_date = datetime.date(TODAY_LOCAL.year, TODAY_LOCAL.month, 1)
    from_date_default = default_date
    form = export_invoices_payments_get_form(default_date, form_type='payments')

    # processing
    if form.process().accepted:
        response.flash = T("Form accepted")

        stream = export_payments_get_export(form.vars.from_date,
                                            form.vars.until_date,
                                            form.vars.invoices_groups_id)

        fname = T("Payments.xlsx")

        response.headers['Content-Type']='application/vnd.ms-excel'
        response.headers['Content-disposition']='attachment; filename=' + fname

        return stream.getvalue()

    # form.process() has to be called before creating a custom form
    # otherwise for hidden formkey fields used for CSRF protection aren't generated
    form_display = export_invoices_payments_get_form_display(form, form_type='payments')
    content = DIV(form_display)

    menu = export_get_menu(request.function)

    back = os_gui.get_button('back', URL('finance', 'invoices'),
                             _class='full-width')

    return dict(content=content,
                back=back,
                menu=menu)


def export_payments_get_export(from_date, until_date, invoices_groups_id):
    """
        Payments export
    """
    # create filestream
    stream = io.BytesIO()

    wb = openpyxl.workbook.Workbook(write_only=True)
    ws = wb.create_sheet(title='Payments')

    header = [
        'Payment Date',
        'Payment Method',
        'Amount',
        'InvoiceID',
        'CustomerID',
        'Customer Name',
        'Date created',
        'Date due',
        'Description',
        'Invoice Status',
        'Invoice Amount',
        'School SubscriptionID',
        'School Subscription Name',
        'Subscription Year',
        'Subscription Month'
    ]

    ws.append(header)

    query = (db.invoices_payments.PaymentDate >= from_date)

    if until_date:
        query &= (db.invoices_payments.PaymentDate <= until_date)

    if invoices_groups_id:
        query &= (db.invoices.invoices_groups_id == invoices_groups_id)

    left = [ db.invoices.on(db.invoices_payments.invoices_id ==
                            db.invoices.id),
             db.invoices_amounts.on(db.invoices_amounts.invoices_id ==
                                    db.invoices_payments.invoices_id),
             db.payment_methods.on(db.invoices_payments.payment_methods_id ==
                                   db.payment_methods.id),
             db.invoices_customers_subscriptions.on(
                 db.invoices_customers_subscriptions.invoices_id ==
                 db.invoices.id),
             db.customers_subscriptions.on(
                db.invoices_customers_subscriptions.customers_subscriptions_id ==
                db.customers_subscriptions.id),
             db.school_subscriptions.on(
                db.customers_subscriptions.school_subscriptions_id ==
                db.school_subscriptions.id),
             db.invoices_customers.on(db.invoices_customers.invoices_id ==
                                      db.invoices.id),
             db.auth_user.on(db.invoices_customers.auth_customer_id == db.auth_user.id),

             ]

    rows = db(query).select(db.invoices_payments.ALL,
                            db.invoices.ALL,
                            db.payment_methods.ALL,
                            db.invoices_amounts.ALL,
                            db.customers_subscriptions.ALL,
                            db.school_subscriptions.Name,
                            db.auth_user.id,
                            db.auth_user.display_name,
                            left=left,
                            orderby=db.invoices_payments.PaymentDate|\
                                    db.invoices.InvoiceID)
    for row in rows:
        line = [
            row.invoices_payments.PaymentDate,
            row.payment_methods.Name,
            row.invoices_payments.Amount,
            row.invoices.InvoiceID,
            row.auth_user.id,
            row.auth_user.display_name,
            row.invoices.DateCreated,
            row.invoices.DateDue,
            row.invoices.Description,
            row.invoices.Status,
            row.invoices_amounts.TotalPriceVAT,
            row.customers_subscriptions.school_subscriptions_id,
            row.school_subscriptions.Name,
            row.invoices.SubscriptionYear,
            row.invoices.SubscriptionMonth,
        ]

        ws.append(line)

    wb.save(stream)


    return stream


def export_get_menu(page):
    """
        Returns submenu for export pages
    """
    pages = [
        [ 'export_invoices', T('Invoices'), URL('export_invoices') ],
        [ 'export_payments', T('Payments'), URL('export_payments') ],

    ]

    return os_gui.get_submenu(pages, page, horizontal=True, htype='tabs')



@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('create', 'invoices'))
def cancel_and_create_credit_invoice():
    """
        Cancel invoice and create credit invoice
    """
    iID = request.vars['iID']

    # Duplicate invoice
    invoice = db.invoices(iID)
    invoice.Status = 'cancelled'
    invoice.update_record()

    new_iID = db.invoices.insert(
        invoices_groups_id = invoice.invoices_groups_id,
        CustomerCompany = invoice.CustomerCompany,
        CustomerName = invoice.CustomerName,
        CustomerAddress = invoice.CustomerAddress,
        CustomerListName = invoice.CustomerListName,
        SubscriptionMonth = invoice.SubscriptionMonth,
        SubscriptionYear = invoice.SubscriptionYear,
        Description = invoice.Description,
        Status = 'sent',
        Terms = invoice.Terms,
        Footer = invoice.Footer,
        Note = invoice.Note,
        credit_invoice_for = invoice.id
    )

    # Set InvoiceID and Due date
    new_invoice = Invoice(new_iID)

    # Link credit invoice to customer
    current_invoice = Invoice(iID)
    linked_customer_id = current_invoice.get_linked_customer_id()

    if linked_customer_id:
        new_invoice.link_to_customer(linked_customer_id)

    # Duplicate invoice items with negative prices
    query = (db.invoices_items.invoices_id == invoice.id)
    rows = db(query).select(db.invoices_items.ALL)
    for row in rows:
        if not row.TotalPriceVAT is None:
            totalpricevat = row.TotalPriceVAT * -1
        else:
            totalpricevat = None

        if not row.VAT is None:
            vat = row.VAT * -1
        else:
            vat = None

        if not row.TotalPrice is None:
            totalprice = row.TotalPrice * -1
        else:
            totalprice = None

        db.invoices_items.insert(
            invoices_id = new_iID,
            Sorting = row.Sorting,
            ProductName = row.ProductName,
            Description = row.Description,
            Quantity = row.Quantity,
            Price = row.Price * -1,
            tax_rates_id = row.tax_rates_id,
            TotalPriceVAT = totalpricevat,
            VAT = vat,
            TotalPrice = totalprice,
            accounting_glaccounts_id = row.accounting_glaccounts_id,
            accounting_costcenters_id = row.accounting_costcenters_id
        )

    new_invoice.set_amounts()

    session.flash = T('You are now editing the credit invoice for invoice ') + invoice.InvoiceID

    redirect(URL('invoices', 'edit', vars={'iID':new_iID}))


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'invoices'))
def open_on_date():
    """
    List invoices not paid on a given date
    """
    from openstudio.os_reports import Reports

    response.title = T("Invoices")
    response.subtitle = T("Open on date")
    open_on_date_process_request_vars()

    reports = Reports()
    rows = reports.get_invoices_open_on_date(session.invoices_open_on_date_date)

    header = THEAD(TR(
        TH(T('Status')),
        TH(T('Invoice ID')),
        TH(T('Invoice Date')),
        TH(T('Customer')),
        TH(T('Amount')),
        TH(T('Amount paid on'), ' ',
           session.invoices_open_on_date_date.strftime(DATE_FORMAT)),
        TH(T("Balance"))
    ))

    balance_total = 0
    table = TABLE(header, _class="table table-striped table-hover")
    for i, row in enumerate(rows):
        repr_row = list(rows[i:i + 1].render())[0]

        balance = row.invoices_amounts.Balance
        if not balance:
            balance = row.invoices_amounts.TotalPriceVAT

        tr = TR(
            TD(repr_row.invoices.Status),
            TD(A(row.invoices.InvoiceID,
                 _href=URL('invoices', 'edit', vars={'iID': row.invoices.id}))),
            TD(repr_row.invoices.DateCreated),
            TD(A(row.auth_user.display_name,
                 _href=URL('customers', 'edit', args=[row.auth_user.id]))),
            TD(repr_row.invoices_amounts.TotalPriceVAT),
            TD(repr_row.invoices_amounts.Paid),
            TD(represent_decimal_as_amount(balance)),
        )

        if balance:
            balance_total += balance

        table.append(tr)


    result = open_on_date_get_form(session.invoices_open_on_date_date)
    form = result['form']
    content_top = result['form_display']

    content_top.append(DIV(SPAN(
        LABEL(T("Total balance on"), ' ',
              session.invoices_open_on_date_date.strftime(DATE_FORMAT)), BR(),
        represent_decimal_as_amount(balance_total),
        _class='pull-right'),
    _class='col-md-9'
    ))

    today = os_gui.get_button(
        'noicon',
        URL('open_on_date_today'),
        title = T("Today"),
        btn_size = "",
        _class="pull-right"
    )

    export = open_on_date_get_export()

    back = os_gui.get_button(
        'back',
        URL('finance', 'invoices')
    )

    return dict(
        form = content_top,
        content = table,
        export = export,
        submit = SPAN(result['submit'], _class='pull-right'),
        header_tools = today,
        back = back,
    )


def open_on_date_get_export(var=None):
    """
        Returns dict with export button and bs3 modal containing the links
        to different export options.
    """
    export = ''

    if auth.has_membership(group_id='Admins') or auth.has_permission('read', 'reports_tax_summary'):
        open_on_date = A((os_gui.get_fa_icon('fa-table'),
                          T("Excel export")),
                          _href=URL('open_on_date_export'),
                          _class='textalign_left')

        links = [
            open_on_date
        ]

        export = os_gui.get_dropdown_menu(
            links=links,
            btn_text='',
            btn_icon='download',
            btn_size='btn',
            menu_class='pull-right')

    return export


def open_on_date_process_request_vars(var=None):
    """
        This function takes the request.vars as a argument and
    """
    from general_helpers import get_last_day_month
    from general_helpers import datestr_to_python

    today = TODAY_LOCAL
    if 'date' in request.vars:
        date = datestr_to_python(DATE_FORMAT, request.vars['date'])
    elif not session.invoices_open_on_date_date is None:
        date = session.invoices_open_on_date_date
    else:
        date = today

    session.invoices_open_on_date_date = date


def open_on_date_get_form(date):
    """
    Get form for open_on_date page
    :param date: datetime.date
    :return: form
    """
    from general_helpers import set_form_id_and_get_submit_button

    form = SQLFORM.factory(
        Field('date', 'date', required=True,
            default=date,
            requires=IS_DATE_IN_RANGE(format=DATE_FORMAT,
                                      minimum=datetime.date(1900,1,1),
                                      maximum=datetime.date(2999,1,1)),
            represent=represent_date,
            label=T("Date to show open invoices"),
            widget=os_datepicker_widget),
        # Field('school_locations_id', db.school_locations,
        #       requires=IS_IN_DB(db(loc_query),
        #                         'school_locations.id',
        #                         '%(Name)s',
        #                         zero=T("All locations")),
        #       default=session.reports_tax_summary_index_school_locations_id,
        #       represent=lambda value, row: locations_dict.get(value, T("No location")),
        #       label=T("Location")),
        formstyle='bootstrap3_stacked',
        submit_button=T("Run report")
    )

    result = set_form_id_and_get_submit_button(form, 'MainForm')
    form = result['form']
    submit = result['submit']

    form_display = DIV(
        XML('<form id="MainForm" action="#" enctype="multipart/form-data" method="post">'),
        DIV(LABEL(form.custom.label.date),
            form.custom.widget.date,
            _class='col-md-3'
        ),
        form.custom.end,
        _class='row'
    )

    return dict(
        form=result['form'],
        submit=result['submit'],
        form_display=form_display
    )


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'invoices'))
def open_on_date_today():
    """
    Set date session var for open_on_date to today
    :return: Redirect to open_on_date
    """
    session.invoices_open_on_date_date = TODAY_LOCAL

    redirect(URL('open_on_date'))


@auth.requires(auth.has_membership(group_id='Admins') or \
               auth.has_permission('read', 'invoices'))
def open_on_date_export():
    """
    Export invoices open on given date to Excel
    :return: Excel worksheet
    """
    from openstudio.os_reports import Reports
    reports = Reports()

    date = session.invoices_open_on_date_date
    sheet_title = "Invoices open on " + date.strftime(DATE_FORMAT)

    # create filestream
    stream = io.BytesIO()

    wb = openpyxl.workbook.Workbook(write_only=True)
    # write the sheet for all mail addresses
    ws = wb.create_sheet(title=sheet_title)
    reports = Reports()
    rows = reports.get_invoices_open_on_date(
        session.invoices_open_on_date_date,
    )

    ws.append([
        "Status",
        "Invoice ID",
        "Date created",
        "Customer ID",
        "Customer",
        "Amount",
        "Amount paid on " + date.strftime(DATE_FORMAT),
        "Balance"
    ])

    for i, row in enumerate(rows):
        repr_row = list(rows[i:i + 1].render())[0]

        balance = row.invoices_amounts.Balance
        paid = row.invoices_amounts.Paid or 0
        if not balance:
            balance = row.invoices_amounts.TotalPriceVAT - paid

        ws.append([
            row.invoices.Status,
            row.invoices.InvoiceID,
            row.invoices.DateCreated.strftime(DATE_FORMAT),
            row.auth_user.id,
            row.auth_user.display_name,
            round(row.invoices_amounts.TotalPriceVAT, 2),
            round(paid, 2),
            round(balance, 2)
        ])

    fname = T(sheet_title.replace(' ','_')) + '.xlsx'
    wb.save(stream)

    response.headers['Content-Type'] = 'application/vnd.ms-excel'
    response.headers['Content-disposition'] = 'attachment; filename=' + fname

    return stream.getvalue()